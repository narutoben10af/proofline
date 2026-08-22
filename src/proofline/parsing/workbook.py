from __future__ import annotations

import hashlib
from datetime import date, datetime
from io import BytesIO
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from proofline.contracts import SourceSpan, SpreadsheetSourceRef
from proofline.parsing.models import ExtractedCell, ExtractionMethod, ExtractionWarning

MAX_WORKBOOK_BYTES = 15 * 1024 * 1024
MAX_SHEETS = 50
MAX_CELLS = 20_000
MAX_ARCHIVE_ENTRIES = 1_000
MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024


class WorkbookExtractionError(ValueError):
    pass


class StructuralXlsxAdapter:
    def extract_cells(self, content: bytes, document_id: str) -> tuple[ExtractedCell, ...]:
        if len(content) > MAX_WORKBOOK_BYTES:
            raise WorkbookExtractionError("workbook exceeds the extraction size limit")
        _validate_xlsx_package(content)
        try:
            values = load_workbook(
                BytesIO(content), read_only=True, data_only=True, keep_links=False
            )
            formulas = load_workbook(
                BytesIO(content), read_only=True, data_only=False, keep_links=False
            )
        except Exception as error:
            raise WorkbookExtractionError("workbook could not be opened") from error
        try:
            if len(values.sheetnames) > MAX_SHEETS:
                raise WorkbookExtractionError("workbook exceeds the sheet limit")
            extracted: list[ExtractedCell] = []
            for sheet_name in values.sheetnames:
                value_sheet = values[sheet_name]
                formula_sheet = formulas[sheet_name]
                if value_sheet.max_row * value_sheet.max_column > MAX_CELLS:
                    raise WorkbookExtractionError("workbook exceeds the cell limit")
                for value_row, formula_row in zip(
                    value_sheet.iter_rows(), formula_sheet.iter_rows(), strict=True
                ):
                    for value_cell, formula_cell in zip(value_row, formula_row, strict=True):
                        formula = formula_cell.value if formula_cell.data_type == "f" else None
                        value = value_cell.value
                        if value is None and formula is None:
                            continue
                        if len(extracted) >= MAX_CELLS:
                            raise WorkbookExtractionError("workbook exceeds the cell limit")
                        display = _display_value(value if value is not None else formula)
                        warnings = ()
                        if formula is not None:
                            if value is None:
                                warnings = (
                                    ExtractionWarning(
                                        code="formula_without_cached_value",
                                        message=(
                                            "Formula has no cached result; no calculation was "
                                            "executed."
                                        ),
                                    ),
                                )
                            else:
                                warnings = (
                                    ExtractionWarning(
                                        code="formula_cached_value_unverified",
                                        message=(
                                            "Formula result is a workbook cache; it was not "
                                            "recalculated or verified."
                                        ),
                                    ),
                                )
                        ref = SpreadsheetSourceRef(
                            document_id=document_id,
                            sheet=sheet_name,
                            cell=value_cell.coordinate,
                            display_value=display,
                        )
                        extracted.append(
                            ExtractedCell(
                                span=SourceSpan(
                                    id=_span_id(document_id, sheet_name, value_cell.coordinate),
                                    document_version_id=document_id,
                                    source=ref,
                                ),
                                method=ExtractionMethod.SPREADSHEET,
                                data_type="formula" if formula is not None else _data_type(value),
                                warnings=warnings,
                            )
                        )
            return tuple(extracted)
        finally:
            values.close()
            formulas.close()


def _validate_xlsx_package(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            members = archive.infolist()
            if len(members) > MAX_ARCHIVE_ENTRIES:
                raise WorkbookExtractionError("XLSX package exceeds the archive entry limit")
            if sum(member.file_size for member in members) > MAX_UNCOMPRESSED_BYTES:
                raise WorkbookExtractionError("XLSX package exceeds the uncompressed size limit")
            if any(member.flag_bits & 0x1 for member in members):
                raise WorkbookExtractionError("encrypted XLSX packages are not supported")
            names = [member.filename for member in members]
            if len(names) != len(set(names)):
                raise WorkbookExtractionError("XLSX package contains duplicate entries")
            name_set = set(names)
            if "[Content_Types].xml" not in name_set or "xl/workbook.xml" not in name_set:
                raise WorkbookExtractionError("content is not an XLSX workbook")
            content_types = archive.read("[Content_Types].xml")
    except BadZipFile as error:
        raise WorkbookExtractionError("content is not an XLSX package") from error
    lower_names = tuple(name.lower() for name in names)
    if (
        b"macroenabled" in content_types.lower()
        or any(name.endswith("vbaproject.bin") for name in lower_names)
        or any(name.startswith(("xl/macrosheets/", "xl/dialogsheets/")) for name in lower_names)
    ):
        raise WorkbookExtractionError("macro-enabled workbooks are not supported")


def _span_id(document_id: str, sheet: str, cell: str) -> str:
    location = f"{document_id}\0{sheet}\0{cell}"
    digest = hashlib.sha256(location.encode("utf-8")).hexdigest()[:20]
    return f"span:{document_id}:xlsx:{digest}"


def _display_value(value: object) -> str:
    if isinstance(value, datetime | date):
        return value.isoformat()
    return str(value)[:256]


def _data_type(value: object) -> str:
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int | float):
        return "number"
    if isinstance(value, datetime | date):
        return "date"
    if isinstance(value, str) and value.startswith("#"):
        return "error"
    return "text"
